import streamlit as st
import random
import calendar
import io
import copy
import json
import os
from datetime import date

try:
    import holidays as holidays_lib
    HOLIDAYS_PKG = True
except ImportError:
    HOLIDAYS_PKG = False

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font

# --- 1. 전역 설정 ---
calendar.setfirstweekday(calendar.SUNDAY)
MEMBERS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "members.json")
DEFAULT_MEMBERS = ["양기윤", "전소영", "임채성", "홍부휘", "이지용",
                   "조현진", "정용채", "강창신", "김덕기", "우성대", "홍그린", "강다현"]

# --- 공휴일 조회 ---
HOLIDAY_FALLBACK = {
    2025: {1: [1, 28, 29, 30], 3: [1], 5: [5, 6], 6: [6], 8: [15],
           9: [5, 6, 7, 8], 10: [3, 9], 12: [25]},
    2026: {1: [1], 2: [16, 17, 18], 3: [1, 2], 5: [5, 24, 25],
           6: [6], 8: [15, 17], 9: [24, 25, 26], 10: [3, 5, 9], 12: [25]},
    2027: {1: [1], 2: [7, 8, 9], 3: [1], 5: [5], 6: [6], 8: [15, 16],
           9: [14, 15, 16], 10: [3, 4, 9], 12: [25]},
}

def get_holidays(year, month):
    if HOLIDAYS_PKG:
        kr = holidays_lib.country_holidays('KR', years=year)
        return [d.day for d in kr.keys() if d.year == year and d.month == month]
    return HOLIDAY_FALLBACK.get(year, {}).get(month, [])

# --- 팀원 파일 영속 관리 ---
def load_members():
    if os.path.exists(MEMBERS_FILE):
        with open(MEMBERS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return DEFAULT_MEMBERS.copy()

def save_members(members):
    with open(MEMBERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(members, f, ensure_ascii=False, indent=2)

# --- 2. 세션 상태 초기화 ---
if 'member_list' not in st.session_state:
    st.session_state.member_list = load_members()

def get_members():
    return st.session_state.member_list

REQUIRED_KEYS = {
    'quotas': {}, 'selection_order': [], 'current_picker_idx': 0, 'slots': [],
    'absentees': set(), 'absentee_prefs': {},
    'history': [], 'manual_mode': False, 'admin_selected_member': None,
    'quota_info': None, 'pass_log': "", 'undo_triggered': False
}
for key, default in REQUIRED_KEYS.items():
    if key not in st.session_state:
        st.session_state[key] = copy.deepcopy(default)

# absentee_prefs 멤버 동기화
for name in get_members():
    if name not in st.session_state.absentee_prefs:
        st.session_state.absentee_prefs[name] = ""

if st.session_state.admin_selected_member is None and get_members():
    st.session_state.admin_selected_member = get_members()[0]

# --- 3. 핵심 제어 함수 ---
def save_history():
    snapshot = {
        'slots': copy.deepcopy(st.session_state.slots),
        'quotas': copy.deepcopy(st.session_state.quotas),
        'current_picker_idx': st.session_state.current_picker_idx,
        'pass_log': st.session_state.pass_log
    }
    st.session_state.history.append(snapshot)
    if len(st.session_state.history) > 20:
        st.session_state.history.pop(0)

def find_next_valid_picker():
    if not st.session_state.selection_order:
        return
    for _ in range(len(st.session_state.selection_order)):
        st.session_state.current_picker_idx = (
            st.session_state.current_picker_idx + 1
        ) % len(st.session_state.selection_order)
        curr = st.session_state.selection_order[st.session_state.current_picker_idx]
        if st.session_state.quotas.get(curr, 0) > 0:
            return

def pass_turn(name):
    rem = st.session_state.quotas.get(name, 0)
    if rem <= 0:
        return
    save_history()
    others = [m for m in get_members() if m != name]
    if others:
        dist = [random.choice(others) for _ in range(rem)]
        for t in dist:
            st.session_state.quotas[t] += 1
        summary_d = {x: dist.count(x) for x in set(dist)}
        st.session_state.pass_log = f"🚫 **{name}** 패스 ➔ " + ", ".join(
            [f"**{k}**(+{v}회)" for k, v in summary_d.items()]
        )
    st.session_state.quotas[name] = 0
    st.session_state.undo_triggered = False
    find_next_valid_picker()
    st.rerun()

# --- 4. UI CSS ---
st.set_page_config(page_title="CARE팀 당직 시스템", layout="wide")
st.markdown("""
    <style>
    .stApp { background-color: #0e1117; color: #ffffff; }
    .day-header-box {
        background-color: #1c1e21; color: #ffffff !important;
        text-align: center; font-weight: 900; padding: 12px;
        border-radius: 8px; margin-bottom: 15px; border: 1px solid #495057;
    }
    .date-tag-normal {
        background-color: #495057; color: #ffffff !important;
        padding: 4px 12px; border-radius: 6px; font-weight: 800;
        border: 1px solid #adb5bd; display: inline-block; margin-bottom: 4px;
    }
    .date-tag-holiday {
        background-color: #c92a2a; color: #ffffff !important;
        padding: 4px 12px; border-radius: 6px; font-weight: 800;
        border: 1px solid #ffa8a8; display: inline-block; margin-bottom: 4px;
    }
    div[data-testid="stButton"] button p { color: white !important; font-weight: 700; }
    div[data-testid="stButton"] button[disabled] {
        background-color: #212529 !important; opacity: 1 !important;
        border: 1px solid #343a40 !important;
    }
    .turn-box {
        background-color: #2b2f36; border-left: 8px solid #fd7e14;
        padding: 15px; border-radius: 10px; color: #ffffff; margin-bottom: 15px;
    }
    .absent-badge {
        color: #ff8787; font-weight: bold; background-color: #c92a2a33;
        padding: 2px 6px; border-radius: 4px; border: 1px solid #c92a2a; margin-left: 5px;
    }
    </style>
""", unsafe_allow_html=True)

# --- 5. 사이드바 ---
with st.sidebar:
    st.title("⚙️ 당직 설정")

    today = date.today()
    col_y, col_m = st.columns(2)
    sel_year = col_y.number_input("연도", 2025, 2030, today.year)
    sel_month = col_m.number_input("월", 1, 12, today.month)

    if st.button("📅 달력 초기화 (새 달 시작)", use_container_width=True):
        cal = calendar.monthcalendar(sel_year, sel_month)
        h_days = set(get_holidays(sel_year, sel_month))
        new_slots = []
        slot_id = 0
        for week in cal:
            for c_idx, day in enumerate(week):
                if day == 0:
                    continue
                is_h = (c_idx == 0 or c_idx == 6 or day in h_days)
                if is_h:
                    new_slots.append({"day": day, "type": "Day", "owner": None,
                                      "id": slot_id, "is_heavy": True})
                    slot_id += 1
                new_slots.append({"day": day, "type": "Night", "owner": None,
                                  "id": slot_id, "is_heavy": is_h})
                slot_id += 1
        st.session_state.update({
            'slots': new_slots, 'quotas': {}, 'selection_order': [],
            'current_picker_idx': 0, 'history': [], 'pass_log': "",
            'quota_info': None, 'undo_triggered': False
        })
        st.rerun()

    st.divider()
    st.session_state.manual_mode = st.toggle("🛡️ 수동 모드 (순번 무시)")
    if st.session_state.manual_mode and get_members():
        st.session_state.admin_selected_member = st.selectbox("배정 대상 선택", get_members())

    st.divider()

    # ── 팀원 관리 (CRUD) ──
    st.subheader("👥 팀원 관리")
    add_col, btn_col = st.columns([3, 1])
    new_name = add_col.text_input(
        "팀원 이름", key="new_member_input",
        label_visibility="collapsed", placeholder="새 팀원 이름"
    )
    if btn_col.button("➕", help="팀원 추가"):
        n = new_name.strip()
        if n and n not in st.session_state.member_list:
            st.session_state.member_list.append(n)
            st.session_state.absentee_prefs[n] = ""
            save_members(st.session_state.member_list)
            st.rerun()
        elif n:
            st.warning("이미 있는 이름입니다.")

    for name in list(st.session_state.member_list):
        r1, r2 = st.columns([4, 1])
        r1.write(f"👤 {name}")
        if r2.button("🗑️", key=f"del_{name}", help=f"{name} 삭제"):
            st.session_state.member_list.remove(name)
            save_members(st.session_state.member_list)
            st.rerun()

    st.divider()

    # ── 개인 설정 ──
    st.subheader("📋 개인 설정")
    for name in sorted(get_members()):
        with st.expander(f"⚙️ {name}"):
            is_abs = st.checkbox("부재중 체크", key=f"abs_{name}",
                                  value=(name in st.session_state.absentees))
            if is_abs:
                st.session_state.absentees.add(name)
            else:
                st.session_state.absentees.discard(name)
            st.session_state.absentee_prefs[name] = st.text_input(
                "희망 ID(쉼표)", value=st.session_state.absentee_prefs.get(name, ""),
                key=f"p_{name}"
            )

# --- 6. 메인 화면 ---
st.title(f"📅 {sel_year}년 {sel_month}월 당직 배정")

members = get_members()
n_members = len(members)

col_info, col_cal = st.columns([1, 2.3])

with col_info:
    st.subheader("🎲 추첨 및 순위 조정")

    if st.button("🔢 1. 근무 횟수 추첨", use_container_width=True, disabled=(n_members == 0)):
        t = len(st.session_state.slots)
        b, e = divmod(t, n_members)
        tmp = members.copy()
        random.shuffle(tmp)
        h, l = sorted(tmp[:e]), sorted(tmp[e:])
        st.session_state.quotas = {n: b + 1 if n in h else b for n in members}
        st.session_state.quota_info = (b + 1, h, b, l)

    rank_col1, rank_col2 = st.columns(2)
    if rank_col1.button("🏃 2-A. 랜덤 순위", use_container_width=True):
        st.session_state.selection_order = random.sample(members, n_members)
        st.session_state.current_picker_idx = 0
        st.session_state.undo_triggered = False
        st.success("랜덤 순위 완료!")

    with st.expander("🏃 2-B. 순위 수동 조정"):
        manual_order = st.multiselect(
            "순서대로 선택", members,
            default=st.session_state.selection_order if st.session_state.selection_order else []
        )
        if st.button("✅ 수동 순위 적용"):
            if len(manual_order) == n_members:
                st.session_state.selection_order = manual_order
                st.session_state.current_picker_idx = 0
                st.session_state.undo_triggered = False
                st.success("완료!")
                st.rerun()
            else:
                st.error(f"{n_members}명 모두 선택해야 합니다. (현재 {len(manual_order)}명)")

    if st.session_state.quota_info:
        b1, h1, b2, l2 = st.session_state.quota_info
        st.info(f"📍 **{b1}회**: {', '.join(h1)}\n\n📍 **{b2}회**: {', '.join(l2)}")

    st.divider()
    ctrl1, ctrl2 = st.columns(2)
    if ctrl1.button("↩️ 되돌리기", use_container_width=True,
                    disabled=not st.session_state.history):
        if st.session_state.history:
            last = st.session_state.history.pop()
            st.session_state.update({
                'slots': last['slots'], 'quotas': last['quotas'],
                'current_picker_idx': last['current_picker_idx'],
                'pass_log': last['pass_log'], 'undo_triggered': True
            })
            st.rerun()

    if ctrl2.button("🚫 패스(배분)", use_container_width=True):
        if st.session_state.selection_order:
            pass_turn(st.session_state.selection_order[st.session_state.current_picker_idx])

    if st.session_state.pass_log:
        st.warning(st.session_state.pass_log)

    st.subheader("📋 순위별 대기열")
    if st.session_state.selection_order:
        curr_p = st.session_state.selection_order[st.session_state.current_picker_idx]
        if st.session_state.quotas.get(curr_p, 0) <= 0:
            find_next_valid_picker()

        for idx, name in enumerate(st.session_state.selection_order):
            q = st.session_state.quotas.get(name, 0)
            if q <= 0:
                continue
            raw_prefs = [
                x.strip() for x in
                st.session_state.absentee_prefs.get(name, "").split(',')
                if x.strip().isdigit()
            ]
            rem_prefs = [
                p for p in raw_prefs
                if int(p) < len(st.session_state.slots)
                and st.session_state.slots[int(p)]['owner'] is None
            ]
            is_turn = (idx == st.session_state.current_picker_idx)
            rank_label = f"{idx + 1}위: {name}"
            abs_tag = '<span class="absent-badge">[부재중]</span>' if name in st.session_state.absentees else ""
            pref_txt = f" | 🌟 남음: {', '.join(rem_prefs)}" if rem_prefs else ""

            if is_turn:
                st.markdown(
                    f'<div class="turn-box"><b>👉 {rank_label}{abs_tag} ({q}회){pref_txt}</b></div>',
                    unsafe_allow_html=True
                )
                if name in st.session_state.absentees and q > 0:
                    if st.session_state.undo_triggered:
                        st.info("↩️ 자동 배정 일시 정지됨")
                        if st.button("자동 배정 재개"):
                            st.session_state.undo_triggered = False
                            st.rerun()
                    else:
                        if rem_prefs:
                            target_id = int(rem_prefs[0])
                            save_history()
                            st.session_state.slots[target_id]['owner'] = name
                            st.session_state.quotas[name] -= 1
                            find_next_valid_picker()
                            st.rerun()
                        else:
                            pass_turn(name)
            else:
                st.markdown(f"• {rank_label}{abs_tag} ({q}회){pref_txt}", unsafe_allow_html=True)

with col_cal:
    h_cols = st.columns(7)
    days_kr = ["일", "월", "화", "수", "목", "금", "토"]
    for i, h in enumerate(days_kr):
        h_cols[i].markdown(f'<div class="day-header-box">{h}</div>', unsafe_allow_html=True)

    if st.session_state.slots:
        cal_grid = calendar.monthcalendar(sel_year, sel_month)
        h_days = get_holidays(sel_year, sel_month)
        for week in cal_grid:
            w_cols = st.columns(7)
            for i, day in enumerate(week):
                if day == 0:
                    continue
                is_h = (i == 0 or i == 6 or day in h_days)
                tag_class = "date-tag-holiday" if is_h else "date-tag-normal"
                with w_cols[i]:
                    st.markdown(f'<div class="{tag_class}">{day}일</div>', unsafe_allow_html=True)
                    for s in [sl for sl in st.session_state.slots if sl['day'] == day]:
                        slot_icon = "🌅 주간" if s['type'] == 'Day' else "🌙 야간"
                        if s['owner']:
                            st.button(
                                f"👤 {s['owner']}", key=f"b{s['id']}",
                                disabled=True, use_container_width=True
                            )
                        else:
                            if st.button(slot_icon, key=f"b{s['id']}", use_container_width=True):
                                save_history()
                                st.session_state.undo_triggered = False
                                if st.session_state.manual_mode:
                                    target = st.session_state.admin_selected_member
                                elif st.session_state.selection_order:
                                    target = st.session_state.selection_order[
                                        st.session_state.current_picker_idx
                                    ]
                                else:
                                    target = None
                                if target and (
                                    st.session_state.quotas.get(target, 0) > 0
                                    or st.session_state.manual_mode
                                ):
                                    s['owner'] = target
                                    st.session_state.quotas[target] -= 1
                                    if not st.session_state.manual_mode:
                                        find_next_valid_picker()
                                    st.rerun()

# --- 7. 당직 현황 요약표 ---
if st.session_state.slots:
    st.divider()
    st.subheader("📊 당직 현황 요약")

    duty_summary = {name: {"주간": 0, "야간": 0} for name in members}
    total_slots = len(st.session_state.slots)
    assigned_count = 0

    for s in st.session_state.slots:
        if s['owner']:
            assigned_count += 1
            if s['owner'] in duty_summary:
                key = "주간" if s['type'] == 'Day' else "야간"
                duty_summary[s['owner']][key] += 1

    prog_pct = assigned_count / total_slots if total_slots else 0
    st.progress(prog_pct, text=f"배정 진행률: {assigned_count}/{total_slots} ({prog_pct * 100:.1f}%)")

    # 4명씩 한 줄로 표시
    chunk_size = 4
    for row_start in range(0, len(members), chunk_size):
        row_members = members[row_start:row_start + chunk_size]
        cols = st.columns(chunk_size)
        for ci, name in enumerate(row_members):
            v = duty_summary[name]
            total = v['주간'] + v['야간']
            cols[ci].metric(
                label=name,
                value=f"총 {total}회",
                delta=f"주간 {v['주간']} / 야간 {v['야간']}"
            )

# --- 8. 엑셀 저장 ---
def make_excel():
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = f"{sel_year}.{sel_month}월"
    headers = ["일", "월", "화", "수", "목", "금", "토"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = PatternFill("solid", fgColor="333333")
        cell.font = Font(color="FFFFFF", bold=True)
        ws.column_dimensions[cell.column_letter].width = 18

    day_map = {d: {"Day": "", "Night": ""} for d in range(1, 32)}
    for s in st.session_state.slots:
        if s['owner']:
            day_map[s['day']][s['type']] = s['owner']

    cal_grid = calendar.monthcalendar(sel_year, sel_month)
    h_days_xl = get_holidays(sel_year, sel_month)
    for r_idx, week in enumerate(cal_grid, 2):
        ws.row_dimensions[r_idx].height = 60
        for c_idx, day in enumerate(week):
            if day == 0:
                continue
            day_txt = day_map[day]['Day']
            night_txt = day_map[day]['Night']
            cell_text = f"[{day}일]"
            if day_txt:
                cell_text += f"\n주: {day_txt}"
            if night_txt:
                cell_text += f"\n야: {night_txt}"
            cell = ws.cell(r_idx, c_idx + 1, cell_text)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            if c_idx == 0 or day in h_days_xl:
                cell.fill = PatternFill("solid", fgColor="ffc9c9")
            elif c_idx == 6:
                cell.fill = PatternFill("solid", fgColor="d0ebff")

    # 요약 시트 추가
    ws2 = wb.create_sheet(title="현황요약")
    ws2.append(["이름", "주간 당직", "야간 당직", "합계"])
    for name in members:
        v = duty_summary[name]
        ws2.append([name, v['주간'], v['야간'], v['주간'] + v['야간']])

    wb.save(output)
    return output.getvalue()

st.divider()
if st.session_state.slots:
    st.download_button(
        "💾 당직표 엑셀 저장",
        data=make_excel(),
        file_name=f"CARE팀_{sel_year}_{sel_month:02d}월.xlsx",
        use_container_width=True,
        type="primary"
    )
